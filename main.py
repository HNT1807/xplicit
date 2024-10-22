import streamlit as st
import pandas as pd
import io
import zipfile
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule
from openpyxl import load_workbook, Workbook

def update_search_words():
    st.session_state.custom_words = st.session_state.search_words
    
def highlight_modified_cells(writer, sheet_name, modified_rows, version_col_name='Version'):
    workbook = writer.book
    worksheet = workbook[sheet_name]

    # Find the column index for 'Version' (case-insensitive)
    version_col = None
    for idx, cell in enumerate(worksheet[1], start=1):
        if cell.value and cell.value.lower() == version_col_name.lower():
            version_col = idx
            break

    if version_col:
        # Define the highlighting style
        yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

        # Apply the highlighting to the modified cells
        for row in modified_rows:
            cell = worksheet.cell(row=row, column=version_col)
            cell.fill = yellow_fill
    else:
        print(f"Warning: '{version_col_name}' column not found. Highlighting not applied.")

def create_report(all_reports):
    wb = Workbook()
    ws = wb.active
    ws.title = "Processing Report"

    headers = ["File Name", "Row", "Volume", "Library", "Original Version", "New Version", "Explicit Words Found"]
    ws.append(headers)

    for file_name, report in all_reports.items():
        for item in report:
            try:
                parts = item.split(": ", 1)
                row = parts[0].split(" ")[1]

                # Extract Volume and Library
                volume = "N/A"
                library = "N/A"
                if "Volume:" in item:
                    volume = item.split("Volume: ")[1].split(", Library:")[0]
                if "Library:" in item:
                    library = item.split("Library: ")[1].split(", Original Version:")[0]

                # Extract versions and explicit words
                versions_part = item.split("Original Version: ")[1]
                original_version, rest = versions_part.split("' became '")
                new_version, explicit_words = rest.split(" >>> ")
                original_version = original_version.strip("'")
                new_version = new_version.strip("'")

                ws.append([file_name, row, volume, library, original_version, new_version, explicit_words])
            except Exception as e:
                print(f"Error processing report item: {item}")
                print(f"Error details: {str(e)}")
                # Append a row with available information
                ws.append([file_name, "Error", "Error", "Error", "Error", "Error", str(e)])

    return wb
    
def reset_app():
    # Clear all session state variables
    for key in list(st.session_state.keys()):
        del st.session_state[key]
    # Set a flag to reinitialize the file uploader
    st.session_state.reset_uploader = True
    
def process_version(version, has_explicit_content):
    if has_explicit_content and "Explicit" not in version:
        parts = version.split(', ')
        if parts[0] in ["Full", "Full Mix", "Main"]:
            parts[0] += " Explicit"
        else:
            parts[0] += " Explicit"
        return ", ".join(parts)
    return version  # Return original version if no changes are needed

def highlight_modified_cells(writer, sheet_name, modified_rows, version_col_name='Version'):
    workbook = writer.book
    worksheet = workbook[sheet_name]

    # Find the column index for 'Version' (case-insensitive)
    version_col = None
    for idx, cell in enumerate(worksheet[1], start=1):
        if cell.value and cell.value.lower() == version_col_name.lower():
            version_col = idx
            break

    if version_col:
        # Define the highlighting style
        yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

        # Apply the highlighting to the modified cells
        for row in modified_rows:
            cell = worksheet.cell(row=row, column=version_col)
            cell.fill = yellow_fill
    else:
        print(f"Warning: '{version_col_name}' column not found. Highlighting not applied.")
        
def process_excel(df, search_words):
    report = []
    modified_rows = []

    # Convert column names to lowercase for case-insensitive comparison
    df.columns = df.columns.str.lower()

    required_columns = ['lyrics', 'version', 'volume', 'library']
    if not all(col in df.columns for col in required_columns):
        missing_columns = [col for col in required_columns if col not in df.columns]
        return df, [], [
            f"Error: Required column(s) {', '.join(missing_columns)} not found in the Excel file. Available columns are: {', '.join(df.columns)}"]

    # Define word boundaries (including space, punctuation, and |)
    word_boundaries = r'[\s.,;:!?()\[\]{}|]'

    modified_df = df.copy()
    for index, row in df.iterrows():
        lyrics = str(row['lyrics']).lower()
        version = str(row['version'])
        volume = str(row['volume'])
        library = str(row['library'])

        # Add spaces at the beginning and end of lyrics for boundary checking
        lyrics = f" {lyrics} "
        
        # Find explicit words with proper word boundaries
        found_explicit_words = []
        for word in search_words:
            word = word.lower()
            # Create a regular expression pattern with word boundaries
            import re
            pattern = f"{word_boundaries}{re.escape(word)}{word_boundaries}"
            
            if re.search(pattern, lyrics):
                found_explicit_words.append(word)

        has_explicit_content = bool(found_explicit_words)
        new_version = process_version(version, has_explicit_content)
        
        if new_version != version:
            report.append(
                f"Row {index + 2}: Volume: {volume}, Library: {library}, Original Version: '{version}' became '{new_version}' >>> {found_explicit_words}")
            modified_df.at[index, 'version'] = new_version
            modified_rows.append(index + 2)

    return modified_df, modified_rows, report

def highlight_explicit_cells(writer, sheet_name):
    workbook = writer.book
    worksheet = workbook[sheet_name]

    # Find the column index for 'Version_Grouping'
    version_col = None
    for idx, cell in enumerate(worksheet[1]):
        if cell.value and cell.value.lower() == 'version':
            version_col = idx
            break

    if version_col:
        # Define the highlighting style
        yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        dxf = DifferentialStyle(fill=yellow_fill)
        rule = Rule(type="containsText", operator="containsText", text="Explicit", dxf=dxf)
        rule.formula = [f'NOT(ISERROR(SEARCH("Explicit",{chr(64 + version_col)}2)))']

        # Apply the conditional formatting to the 'Version_Grouping' column
        worksheet.conditional_formatting.add(
            f'{chr(64 + version_col)}2:{chr(64 + version_col)}{worksheet.max_row}', rule)
        
def on_file_upload():
    st.session_state.file_uploaded = True

def main():
    st.markdown(
        "<h1 style='text-align: center; color: white;'><b>XPLICIT</b></h1>",
        unsafe_allow_html=True
    )

    # File uploader
    uploaded_files = st.file_uploader("Choose Excel files", type=["xlsx", "xls"],
                                      accept_multiple_files=True,
                                      key="file_uploader")

    # Define default search words
    default_search_words = ["shit", "bullshit", "shithead", "piss", "fuck", "cunt", "cocksucker", "motherfucker",
                            "tits", "pussy", "asshole", "wog", "wop", "nigger", "kike", "gook", "gypsy", "faggot",
                            "goddamn"]

    # Initialize session state variables
    if 'custom_words' not in st.session_state:
        st.session_state.custom_words = default_search_words.copy()
    if 'search_words' not in st.session_state:
        st.session_state.search_words = st.session_state.custom_words.copy()

    # Display and handle search words
    selected_words = st.multiselect(
        "Default list:",
        options=st.session_state.custom_words,
        default=st.session_state.search_words
    )

    # Update session state based on selection
    st.session_state.search_words = selected_words

    # Add field for new words
    new_words = st.text_input("Enter new words (comma separated):")

    # Button to add new words
    if st.button("Add New Words"):
        if new_words:
            new_word_list = [word.strip() for word in new_words.split(',') if word.strip()]
            st.session_state.custom_words.extend(new_word_list)
            st.session_state.custom_words = list(set(st.session_state.custom_words))
            st.session_state.search_words.extend(new_word_list)
            st.success(f"Added {len(new_word_list)} new word(s) to the list.")


    # Button to reset search words to default
    if st.button("Reset to Default Words"):
        st.session_state.custom_words = default_search_words.copy()
        st.session_state.search_words = default_search_words.copy()

    if uploaded_files and st.button("CHECK FOR EXPLICIT WORDS!"):
        all_reports = {}
        processed_files = []
        processing_report = []

        for uploaded_file in uploaded_files:
            try:
                df = pd.read_excel(uploaded_file)
                modified_df, modified_rows, report = process_excel(df, st.session_state.search_words)
                file_report = [f"Processing Report for {uploaded_file.name}:"]
                if not report:
                    file_report.append("No changes were made to the file.")
                else:
                    file_report.extend(report)
                    all_reports[uploaded_file.name] = report

                processing_report.extend(file_report)

                if not any(item.startswith("Error") for item in report):
                    buffer = io.BytesIO()
                    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                        modified_df.to_excel(writer, index=False, sheet_name='Sheet1')
                        highlight_modified_cells(writer, 'Sheet1', modified_rows)
                    processed_files.append((uploaded_file.name, buffer.getvalue()))
            except Exception as e:
                error_message = f"An error occurred processing {uploaded_file.name}: {str(e)}"
                processing_report.append(error_message)
                st.error(error_message)

        # Store results in session state
        st.session_state.processed_files = processed_files
        st.session_state.all_reports = all_reports
        st.session_state.processing_report = processing_report

        # Display processing report if available
    if 'processing_report' in st.session_state and st.session_state.processing_report:
        for item in st.session_state.processing_report:
            st.write(item)

        # Display download buttons if data is available in session state
    if 'processed_files' in st.session_state and st.session_state.processed_files:
        if len(st.session_state.processed_files) == 1:
            st.download_button(
                label=f"DOWNLOAD UPDATED {st.session_state.processed_files[0][0]}",
                data=st.session_state.processed_files[0][1],
                file_name=f"modified_{st.session_state.processed_files[0][0]}",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        else:
            zip_buffer = io.BytesIO()
            with zipfile.ZipFile(zip_buffer, "a", zipfile.ZIP_DEFLATED, False) as zip_file:
                for file_name, file_content in st.session_state.processed_files:
                    zip_file.writestr(f"modified_{file_name}", file_content)

            st.download_button(
                label="DOWNLOAD UPDATED XLS",
                data=zip_buffer.getvalue(),
                file_name="modified_excel_files.zip",
                mime="application/zip"
            )

    if 'all_reports' in st.session_state and st.session_state.all_reports:
        report_wb = create_report(st.session_state.all_reports)
        report_buffer = io.BytesIO()
        report_wb.save(report_buffer)
        report_buffer.seek(0)

        st.download_button(
            label="DOWNLOAD REPORT",
            data=report_buffer,
            file_name="Explicit Report.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        # Add RESET button after all other buttons
    if st.button("RESET"):
        for key in list(st.session_state.keys()):
            del st.session_state[key]
        st.experimental_rerun()

if __name__ == "__main__":
    main()
